"""
Excel export endpoint for monthly KPI reports.
Generates .xlsx with multiple sheets:
- Overview (totals)
- By category
- By machine
- By line
- Top causes
- Detailed event list
"""
import io
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, desc
import uuid

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.core.dependencies import get_current_active_user
from app.models.user import User
from app.models.downtime import DowntimeEvent, DowntimeCategory
from app.models.machine import Machine

router = APIRouter(prefix="/export", tags=["export"])


HEADER_FILL = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CATEGORY_FILL = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _style_header(ws, row_idx: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _format_seconds(secs: int) -> str:
    """Format seconds as Hh Mm."""
    if secs < 60:
        return f"{secs}s"
    h = secs // 3600
    m = (secs % 3600) // 60
    if h > 0:
        return f"{h}h {m}m"
    return f"{m}m"


def _autosize(ws):
    for col in ws.columns:
        max_length = 10
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length, 50)


@router.get("/kpi/monthly.xlsx")
async def export_monthly_kpi(
    year: int = Query(None, ge=2020, le=2030),
    month: int = Query(None, ge=1, le=12),
    line: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export monthly KPI report to Excel format."""
    now = datetime.utcnow()
    year = year or now.year
    month = month or now.month

    period_start = datetime(year, month, 1)
    if month == 12:
        period_end = datetime(year + 1, 1, 1)
    else:
        period_end = datetime(year, month + 1, 1)

    # Base filter
    base_filters = [
        DowntimeEvent.started_at >= period_start,
        DowntimeEvent.started_at < period_end,
        DowntimeEvent.resolved_at.is_not(None),
        DowntimeEvent.category.not_in([DowntimeCategory.FREE_SHIFT, DowntimeCategory.WEEKEND]),
    ]
    if line:
        base_filters.append(Machine.line == line)

    # Get all events for the period
    result = await db.execute(
        select(DowntimeEvent, Machine)
        .join(Machine, Machine.id == DowntimeEvent.machine_id)
        .where(and_(*base_filters))
        .order_by(desc(DowntimeEvent.duration_seconds))
    )
    events = [(dt, m) for dt, m in result.all()]

    # Aggregate data
    total_seconds = sum(dt.duration_seconds for dt, _ in events)
    total_events = len(events)
    total_hours = total_seconds / 3600.0
    availability_loss = (total_seconds / (24 * 3600 * 30)) * 100 if total_seconds > 0 else 0

    # By category
    by_category: dict = {}
    for dt, _ in events:
        cat = dt.category.value
        if cat not in by_category:
            by_category[cat] = {"count": 0, "seconds": 0}
        by_category[cat]["count"] += 1
        by_category[cat]["seconds"] += dt.duration_seconds

    # By machine
    by_machine: dict = {}
    for dt, m in events:
        key = (m.code, m.name)
        if key not in by_machine:
            by_machine[key] = {"count": 0, "seconds": 0}
        by_machine[key]["count"] += 1
        by_machine[key]["seconds"] += dt.duration_seconds

    # By line
    by_line: dict = {}
    for dt, m in events:
        ln = m.line or "Unknown"
        if ln not in by_line:
            by_line[ln] = {"count": 0, "seconds": 0}
        by_line[ln]["count"] += 1
        by_line[ln]["seconds"] += dt.duration_seconds

    # Top causes (sub_category)
    by_cause: dict = {}
    for dt, _ in events:
        sub = dt.sub_category or "(unknown)"
        if sub not in by_cause:
            by_cause[sub] = {"count": 0, "seconds": 0}
        by_cause[sub]["count"] += 1
        by_cause[sub]["seconds"] += dt.duration_seconds

    # Build workbook
    wb = Workbook()

    # --- Sheet 1: Overview ---
    ws = wb.active
    ws.title = "Overview"
    ws.append([f"KPI Report - {period_start.strftime('%B %Y')}"])
    ws[1][0].font = Font(bold=True, size=16, color="0EA5E9")
    ws.append([])
    ws.append(["Period:", f"{period_start.strftime('%d.%m.%Y')} - {(period_end - __import__('datetime').timedelta(days=1)).strftime('%d.%m.%Y')}"])
    ws.append(["Line:", line or "All lines"])
    ws.append(["Generated:", datetime.utcnow().strftime("%d.%m.%Y %H:%M:%S UTC")])
    ws.append(["User:", current_user.full_name])
    ws.append([])
    ws.append(["METRIC", "VALUE"])
    _style_header(ws, ws.max_row, 2)
    rows = [
        ("Total events", total_events),
        ("Total downtime", _format_seconds(total_seconds)),
        ("Total hours", round(total_hours, 2)),
        ("Average duration", _format_seconds(total_seconds // total_events) if total_events else "0s"),
        ("Availability loss (%)", f"{availability_loss:.2f}%"),
    ]
    for r in rows:
        ws.append(r)
    _autosize(ws)

    # --- Sheet 2: By Category ---
    ws = wb.create_sheet("By Category")
    ws.append(["Category", "Events", "Total time", "Total (hours)", "Percentage"])
    _style_header(ws, 1, 5)
    for cat, data in sorted(by_category.items(), key=lambda x: -x[1]["seconds"]):
        pct = (data["seconds"] / total_seconds * 100) if total_seconds > 0 else 0
        ws.append([
            cat,
            data["count"],
            _format_seconds(data["seconds"]),
            round(data["seconds"] / 3600, 2),
            f"{pct:.1f}%",
        ])
    _autosize(ws)

    # --- Sheet 3: By Machine ---
    ws = wb.create_sheet("By Machine")
    ws.append(["Machine (code)", "Name", "Events", "Total time", "Total (hours)", "Percentage"])
    _style_header(ws, 1, 6)
    for (code, name), data in sorted(by_machine.items(), key=lambda x: -x[1]["seconds"]):
        pct = (data["seconds"] / total_seconds * 100) if total_seconds > 0 else 0
        ws.append([
            code,
            name,
            data["count"],
            _format_seconds(data["seconds"]),
            round(data["seconds"] / 3600, 2),
            f"{pct:.1f}%",
        ])
    _autosize(ws)

    # --- Sheet 4: By Line ---
    ws = wb.create_sheet("By Line")
    ws.append(["Line", "Events", "Total time", "Total (hours)", "Percentage"])
    _style_header(ws, 1, 5)
    for ln, data in sorted(by_line.items(), key=lambda x: -x[1]["seconds"]):
        pct = (data["seconds"] / total_seconds * 100) if total_seconds > 0 else 0
        ws.append([
            ln,
            data["count"],
            _format_seconds(data["seconds"]),
            round(data["seconds"] / 3600, 2),
            f"{pct:.1f}%",
        ])
    _autosize(ws)

    # --- Sheet 5: Top Causes ---
    ws = wb.create_sheet("Top Causes")
    ws.append(["Subcategory (cause)", "Events", "Total time", "Total (hours)", "Percentage"])
    _style_header(ws, 1, 5)
    for cause, data in sorted(by_cause.items(), key=lambda x: -x[1]["seconds"])[:15]:
        pct = (data["seconds"] / total_seconds * 100) if total_seconds > 0 else 0
        ws.append([
            cause,
            data["count"],
            _format_seconds(data["seconds"]),
            round(data["seconds"] / 3600, 2),
            f"{pct:.1f}%",
        ])
    _autosize(ws)

    # --- Sheet 6: Detailed List ---
    ws = wb.create_sheet("Detailed List")
    headers = ["ID", "Machine", "Line", "Category", "Subcategory", "Description",
               "Started", "Closed", "Duration", "Closed by"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    # Load user names for closed_by
    from app.models.user import User as UserModel
    user_ids = {dt.closed_by_user_id for dt, _ in events if dt.closed_by_user_id}
    user_map = {}
    if user_ids:
        ures = await db.execute(select(UserModel).where(UserModel.id.in_(user_ids)))
        for u in ures.scalars().all():
            user_map[u.id] = u.full_name

    for dt, m in events:
        ws.append([
            str(dt.id)[:8].upper(),
            m.code,
            m.line or "",
            dt.category.value,
            dt.sub_category or "",
            dt.problem_description or "",
            dt.started_at.strftime("%d.%m.%Y %H:%M"),
            dt.resolved_at.strftime("%d.%m.%Y %H:%M") if dt.resolved_at else "",
            _format_seconds(dt.duration_seconds),
            user_map.get(dt.closed_by_user_id, "") if dt.closed_by_user_id else "",
        ])
    _autosize(ws)

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"kpi_report_{year}_{month:02d}_{line or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
